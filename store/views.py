from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from .models import Store, Product, Booking, ProductImage, AboutUsContent, AboutFeature
from django.contrib import messages
from django.core.mail import EmailMessage, send_mail
from django.template.loader import render_to_string
from django.contrib.sites.shortcuts import get_current_site
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from django.contrib.auth.tokens import default_token_generator
import openpyxl
from django.http import HttpResponse
import json
import math
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth import login
from django.core.paginator import Paginator
from django import forms
from django.contrib.auth.models import User
from .forms import KhachHangRegisterForm

# Tính khoảng cách 
def haversine_distance(lat1, lon1, lat2, lon2):
    R = 6371
    try:
        dlat = math.radians(lat2 - lat1)
        dlon = math.radians(lon2 - lon1)
        a = math.sin(dlat / 2) * math.sin(dlat / 2) + \
            math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * \
            math.sin(dlon / 2) * math.sin(dlon / 2)
        c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
        return round(R * c, 2)
    except:
        return 0

def calculate_fee_logic(distance_km):
    if distance_km < 5: return 0
    elif 5 <= distance_km <= 10: return distance_km * 5000
    else: return distance_km * 10000

# 1. Trang chủ 
def home_shop(request):
    # Lấy toàn bộ váy cưới (sắp xếp mới nhất)
    products_list = Product.objects.all().order_by('-id') 
    
    # === BẮT ĐẦU: XỬ LÝ LỌC TÌM KIẾM ===
    query = request.GET.get('q')
    price_range = request.GET.get('price_range')

    # Lọc theo tên (nếu có gõ chữ)
    if query:
        products_list = products_list.filter(name__icontains=query)

    # Lọc theo giá (nếu có chọn khoảng giá)
    if price_range == 'under_5':
        products_list = products_list.filter(price__lt=5000000)
    elif price_range == '5_to_10':
        products_list = products_list.filter(price__gte=5000000, price__lte=10000000)
    elif price_range == 'over_10':
        products_list = products_list.filter(price__gt=10000000)
    # === KẾT THÚC: XỬ LÝ LỌC TÌM KIẾM ===

    # Giữ nguyên code chia trang của bạn
    paginator = Paginator(products_list, 4) 
    page_number = request.GET.get('page')
    products = paginator.get_page(page_number)
    
    return render(request, 'home.html', {'products': products})

# TRANG GIỚI THIỆU MỚI (Lấy dữ liệu từ DB)
def about(request):
    about_content = AboutUsContent.objects.first()
    about_features = AboutFeature.objects.all()
    context = {
        'about_content': about_content,
        'about_features': about_features
    }
    return render(request, 'about.html', context)

# 2. Trang bản đồ 
def map_view(request):
    products = Product.objects.all()
    stores = Store.objects.all()
    user_lat = 10.7721 
    user_lon = 106.6983
    is_default = True  
    
    if request.method == 'POST':
        try:
            lat_str = str(request.POST.get('lat', '')).replace(',', '.')
            lon_str = str(request.POST.get('lon', '')).replace(',', '.')
            
            user_lat = float(lat_str)
            user_lon = float(lon_str)
            is_default = False 
        except (ValueError, TypeError):
            pass 

    store_list = []
    for s in stores:
        if not is_default:
            dist = haversine_distance(user_lat, user_lon, s.latitude, s.longitude)
            fee = calculate_fee_logic(dist)
            formatted_fee = "{:,.0f} đ".format(fee).replace(",", ".") if fee > 0 else "Miễn phí"
        else:
            dist = 0
            fee = 0
            formatted_fee = "---" 
        store_list.append({
            'id': s.id,
            'name': s.name, 'address': s.address, 'phone': s.phone,
            'lat': s.latitude, 'lon': s.longitude, 'distance': dist,
            'fee': fee, 'formatted_fee': formatted_fee,
            'image': s.image.url if s.image else 'https://dummyimage.com/400x150/fff0f3/d88a9a&text=Bridal+Luxury'        })
    
    if not is_default:
        store_list.sort(key=lambda x: x['distance'])
        
    stores_json = json.dumps(store_list)
    
    return render(request, 'map.html', {
        'stores_json': stores_json, 'stores': store_list,
        'products': products, 'user_lat': user_lat, 'user_lon': user_lon,
        'is_default': is_default 
    })

# 3. Trang chi tiết sản phẩm 
def product_detail(request, product_id):
    product = get_object_or_404(Product, pk=product_id)
    related_products = Product.objects.filter(store=product.store).exclude(id=product.id)

    if request.method == 'POST':
        if not request.user.is_authenticated:
            messages.warning(request, "Vui lòng đăng nhập để hệ thống giữ chỗ thử váy cho bạn nhé!")
            return redirect('login')
            
        name = request.POST.get('customer_name')
        phone_input = request.POST.get('phone_number') 
        date = request.POST.get('booking_date')
        note = request.POST.get('note')

        is_shipping = request.POST.get('is_shipping') == 'on' 
        address = request.POST.get('address')
        lat = request.POST.get('latitude')
        lon = request.POST.get('longitude')
        fee = request.POST.get('shipping_fee', 0)

        selected_product_ids = request.POST.getlist('product_ids')
        if not selected_product_ids:
            selected_product_ids = [product.id]

        if name and phone_input and date:
            new_booking = Booking.objects.create(
                customer_name=name, 
                phone=phone_input,  
                booking_date=date, 
                note=note
            )

            ship_info_email = "Đến thử trực tiếp tại cửa hàng"
            if is_shipping and address and lat and lon:
                new_booking.address = address
                new_booking.latitude = lat
                new_booking.longitude = lon
                new_booking.shipping_fee = fee
                new_booking.save()
                ship_info_email = f"Giao váy tận nhà\n- Địa chỉ: {address}\n- Phí ship dự kiến: {fee} VNĐ"

            for p_id in selected_product_ids:
                new_booking.products.add(p_id)

            tieu_de = f"[Bridal Luxury] Đơn mới từ: {name}"
            noi_dung = (
                f"Hệ thống vừa nhận 1 đơn đặt lịch mới:\n\n"
                f"- Khách hàng: {name}\n"
                f"- Số điện thoại: {phone_input}\n"
                f"- Ngày hẹn: {date}\n"
                f"- Số lượng váy thử: {len(selected_product_ids)} váy\n"
                f"- Hình thức: {ship_info_email}\n"
                f"- Ghi chú: {note}\n\n"
                f"Vui lòng vào trang Quản lý để duyệt đơn!"
            )
            
            try:
                send_mail(
                    subject=tieu_de,
                    message=noi_dung,
                    from_email='hethong@bridalluxury.com',
                    recipient_list=['admin_test@gmail.com'],
                    fail_silently=True, 
                )
            except Exception as e:
                print(f"Lỗi gửi mail: {e}") 

            messages.success(request, f"✨ Chúc mừng {name}! Đã đặt lịch thành công.")
            return redirect('product_detail', product_id=product.id)
        else:
            messages.error(request, "Vui lòng điền đầy đủ thông tin.")

    return render(request, 'product_detail.html', {'product': product, 'related_products': related_products})

# 4. API tính phí  
def api_calculate_shipping(request):
    store_id = request.GET.get('store_id')
    user_lat = request.GET.get('lat')
    user_lon = request.GET.get('lon')
    if not store_id or not user_lat: return JsonResponse({'error': 'Thiếu dữ liệu'}, status=400)
    try:
        store = Store.objects.get(pk=store_id)
        dist = haversine_distance(store.latitude, store.longitude, float(user_lat), float(user_lon))
        fee = calculate_fee_logic(dist)
        return JsonResponse({'store': store.name, 'distance_km': dist, 'shipping_fee': fee, 'formatted_fee': "{:,.0f}".format(fee).replace(",", ".")})
    except: return JsonResponse({'error': 'Lỗi'}, status=400)

# 5. Hệ thống cửa hàng
def store_list(request):
    store_queryset = Store.objects.all().order_by('-id')
    paginator = Paginator(store_queryset, 6) 
    page_number = request.GET.get('page')
    stores = paginator.get_page(page_number)
    return render(request, 'store_list.html', {'stores': stores})

def store_detail_view(request, store_id):
    store = get_object_or_404(Store, pk=store_id)
    products = Product.objects.filter(store=store).order_by('-id')
    return render(request, 'store_detail_view.html', {'store': store, 'products': products})

# 6. Trang liên hệ 
def contact(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        email = request.POST.get('email', 'Không cung cấp') 
        phone = request.POST.get('phone', 'Không cung cấp') 
        message_content = request.POST.get('message', 'Không có nội dung')

        tieu_de = f"[Bridal Luxury] Yêu cầu hỗ trợ từ: {name}"
        noi_dung = (
            f"Xin chào Admin,\n\n"
            f"Hệ thống vừa nhận được một lời nhắn mới từ trang Liên hệ:\n"
            f"-----------------------------------------\n"
            f"👤 Tên khách hàng: {name}\n"
            f"📞 Số điện thoại: {phone}\n"
            f"📧 Email: {email}\n"
            f"💬 Nội dung lời nhắn:\n{message_content}\n"
            f"-----------------------------------------\n\n"
            f"Admin vui lòng phản hồi khách hàng trong thời gian sớm nhất nhé!"
        )

        try:
            send_mail(
                subject=tieu_de,
                message=noi_dung,
                from_email='contact@bridalluxury.com',
                recipient_list=['admin_test@gmail.com'],
                fail_silently=True, 
            )
        except Exception as e:
            print(f"Lỗi hệ thống khi gửi mail liên hệ: {e}")

        messages.success(request, f"💖 Cảm ơn {name}! Bridal Luxury đã nhận được tin nhắn và sẽ liên hệ lại với bạn sớm nhất.")
        return redirect('contact')
        
    return render(request, 'contact.html')

# 7. Chức năng tài khoản - đăng ký 
def register_user(request):
    if request.user.is_authenticated:
        return redirect('home')

    if request.method == 'POST':
        form = KhachHangRegisterForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data.get('email')
            if User.objects.filter(email=email).exists():
                messages.error(request, "❌ Email này đã được đăng ký! Vui lòng sử dụng email khác.")
                return render(request, 'register.html', {'form': form})

            user = form.save(commit=False) 
            user.is_active = False         
            user.save()                    
            
            username = form.cleaned_data.get('username')

            current_site = get_current_site(request)
            mail_subject = 'Kích hoạt tài khoản Bridal Luxury'
            message = render_to_string('acc_active_email.html', {
                'user': user,
                'domain': current_site.domain,
                'uid': urlsafe_base64_encode(force_bytes(user.pk)),
                'token': default_token_generator.make_token(user),
            })
            try:
                email_send = EmailMessage(mail_subject, message, to=[user.email])
                email_send.content_subtype = "html"
                email_send.send()
            except Exception as e:
                print(f"Lỗi gửi mail cho khách: {e}")

            tieu_de = f"[Bridal Luxury] Thành viên mới gia nhập: {username}"
            noi_dung = (
                f"Xin chào Admin,\n\n"
                f"Hệ thống vừa ghi nhận một thành viên mới đăng ký tài khoản (Đang chờ kích hoạt email):\n"
                f"- Tên tài khoản: {username}\n"
                f"- Email: {user.email}\n"
                f"- Ngày gia nhập: {user.date_joined.strftime('%d/%m/%Y %H:%M')}\n\n"
                f"Vui lòng kiểm tra danh sách người dùng trong trang quản trị."
            )
            
            try:
                send_mail(
                    subject=tieu_de,
                    message=noi_dung,
                    from_email='system@bridalluxury.com',
                    recipient_list=['admin_test@gmail.com'], 
                    fail_silently=True, 
                )
            except Exception as e:
                print(f"Lỗi gửi mail cho admin: {e}")

            messages.success(request, f"✨ Chào mừng {username}! Vui lòng kiểm tra hộp thư {user.email} để kích hoạt tài khoản.")
            return redirect('login') 
        else:
            messages.error(request, "Đăng ký không thành công. Vui lòng kiểm tra lại thông tin.")
    else:
        form = KhachHangRegisterForm()

    form.fields['username'].help_text = "Nhập tên tài khoản viết liền không dấu (Ví dụ: thuytrang123)"
    return render(request, 'register.html', {'form': form})

# 8. Trang quản lý shop 
def custom_manager(request):
    if not request.user.is_authenticated or not request.user.is_staff:
        messages.error(request, "Cảnh báo: Bạn không có quyền truy cập trang quản trị!")
        return redirect('home')
        
    bookings = Booking.objects.all().order_by('-id') 
    products = Product.objects.all().order_by('-id') 
    stores = Store.objects.all().order_by('-id')     
    
    # KÉO DATA CHO TRANG QUẢN LÝ GIỚI THIỆU
    about_content = AboutUsContent.objects.first()
    about_features = AboutFeature.objects.all()

    users = User.objects.all().order_by('-date_joined')

    context = {
        'bookings': bookings,
        'products': products,  
        'stores': stores,      
        'total_bookings': bookings.count(),
        'total_products': products.count(),
        'total_stores': stores.count(),
        'about_content': about_content,
        'about_features': about_features,
        'users': users,
        'total_users': users.count(),
    }
    return render(request, 'manager.html', context)

# 8.1 Hàm Xóa Đơn Đặt Lịch
def delete_booking(request, id):
    if request.user.is_staff:
        Booking.objects.filter(id=id).delete()
        messages.success(request, "Đã xóa đơn đặt lịch thành công!")
    return redirect('custom_manager')

# 8.2 Hàm Sửa Đơn Đặt Lịch 
def edit_booking(request, id):
    if not request.user.is_authenticated or not request.user.is_staff:
        messages.error(request, "Cảnh báo: Bạn không có quyền!")
        return redirect('home')
    booking = get_object_or_404(Booking, pk=id)
    if request.method == 'POST':
        booking.customer_name = request.POST.get('customer_name')
        booking.phone = request.POST.get('phone')
        booking.note = request.POST.get('note', '')
        booking.status = request.POST.get('status') 
        booking.save() 
        messages.success(request, f"Đã cập nhật trạng thái đơn của {booking.customer_name}!")
        return redirect('custom_manager')

    return render(request, 'edit_booking.html', {'booking': booking})

# 8.3 Hàm Xóa Váy Cưới
def delete_product(request, id):
    if request.user.is_staff:
        Product.objects.filter(id=id).delete()
        messages.success(request, "Đã xóa váy cưới khỏi hệ thống!")
    return redirect('custom_manager')

# 8.4 Hàm Thêm Váy Cưới Mới
def add_product(request):
    if request.method == 'POST' and request.user.is_staff:
        name = request.POST.get('name')
        price = request.POST.get('price')
        store_id = request.POST.get('store')
        quantity = request.POST.get('quantity', 1) 
        status = request.POST.get('status', 'available')
        short_description = request.POST.get('short_description', '') 
        description = request.POST.get('description', '')             
        image = request.FILES.get('image')
        
        try:
            store = Store.objects.get(pk=store_id)
            product = Product.objects.create(
                name=name, price=price, store=store, quantity=quantity, status=status,
                short_description=short_description, description=description, image=image
            )
            
            album_anh = request.FILES.getlist('album_anh') 
            if album_anh: 
                for img in album_anh:
                    ProductImage.objects.create(product=product, image=img)
            
            messages.success(request, f"Đã thêm váy cưới '{name}' vào kho thành công!")
        except Store.DoesNotExist:
            messages.error(request, "Lỗi: Không tìm thấy chi nhánh này!")
        except Exception as e:
            messages.error(request, f"Lỗi hệ thống khi nhập kho: {str(e)}")
            
    return redirect('custom_manager')

# 8.5 Hàm Sửa Váy Cưới
def edit_product(request, id):
    if not request.user.is_authenticated or not request.user.is_staff:
        return redirect('home')
        
    product = get_object_or_404(Product, pk=id)
    stores = Store.objects.all()
    
    if request.method == 'POST':
        product.name = request.POST.get('name', product.name)
        product.price = request.POST.get('price', product.price)
        product.quantity = request.POST.get('quantity', product.quantity)
        product.status = request.POST.get('status', product.status)
        product.short_description = request.POST.get('short_description', product.short_description)
        product.description = request.POST.get('description', product.description)
        
        store_id = request.POST.get('store')
        if store_id:
            try:
                product.store = Store.objects.get(pk=store_id)
            except Store.DoesNotExist:
                pass 

        image = request.FILES.get('image')
        if image:
            product.image = image

        product.save() 

        album_anh = request.FILES.getlist('album_anh')
        if album_anh:
            for img in album_anh:
                ProductImage.objects.create(product=product, image=img)

        messages.success(request, f"Đã cập nhật váy cưới '{product.name}' thành công!")
        return redirect('custom_manager')

    return render(request, 'edit_product.html', {'product': product, 'stores': stores})

# 8.6 Hàm Xóa Cửa Hàng
def delete_store(request, id):
    if request.user.is_staff:
        Store.objects.filter(id=id).delete()
        messages.success(request, "Đã xóa cửa hàng!")
    return redirect('custom_manager')

# 8.7 Hàm Thêm Cửa Hàng Mới
def add_store(request):
    if request.method == 'POST' and request.user.is_staff:
        name = request.POST.get('name')
        address = request.POST.get('address')
        phone = request.POST.get('phone')
        lat = request.POST.get('lat')
        lon = request.POST.get('lon')
        image = request.FILES.get('image') 
        Store.objects.create(
            name=name, address=address, phone=phone, 
            latitude=lat, longitude=lon, image=image
        )
        messages.success(request, f"Đã thêm cửa hàng {name} thành công!")
    return redirect('custom_manager')

# 8.8 Hàm Sửa Cửa Hàng
def edit_store(request, id):
    if not request.user.is_authenticated or not request.user.is_staff:
        return redirect('home')

    store = get_object_or_404(Store, pk=id)
    if request.method == 'POST':
        store.name = request.POST.get('name')
        store.phone = request.POST.get('phone')
        store.address = request.POST.get('address')
        store.latitude = request.POST.get('lat')
        store.longitude = request.POST.get('lon')
        image = request.FILES.get('image')
        if image:
            store.image = image
        store.save()
        messages.success(request, f"Đã cập nhật cửa hàng '{store.name}' thành công!")
        return redirect('custom_manager')

    return render(request, 'edit_store.html', {'store': store})

# ==========================================================
# CÁC HÀM XỬ LÝ TRANG QUẢN LÝ GIỚI THIỆU
# ==========================================================
def update_about_content(request):
    if request.method == 'POST' and request.user.is_staff:
        about, created = AboutUsContent.objects.get_or_create(id=1) 
        
        about.main_title = request.POST.get('main_title')
        about.slogan = request.POST.get('slogan')
        about.origin_title = request.POST.get('origin_title')
        about.origin_content = request.POST.get('origin_content')
        about.tech_title = request.POST.get('tech_title')
        about.tech_content = request.POST.get('tech_content')
        about.tech_quote = request.POST.get('tech_quote')
        
        if 'origin_image' in request.FILES:
            about.origin_image = request.FILES['origin_image']
        if 'tech_image' in request.FILES:
            about.tech_image = request.FILES['tech_image']
            
        about.save()
        messages.success(request, 'Đã cập nhật nội dung trang Giới Thiệu!')
        
    return redirect('custom_manager')

def add_about_feature(request):
    if request.method == 'POST' and request.user.is_staff:
        title = request.POST.get('title')
        description = request.POST.get('description')
        icon_image = request.FILES.get('icon_image')
        
        AboutFeature.objects.create(
            title=title,
            description=description,
            icon_image=icon_image
        )
        messages.success(request, 'Đã thêm tiêu chí nổi bật mới!')
        
    return redirect('custom_manager')

def delete_about_feature(request, feature_id):
    if request.user.is_staff:
        feature = get_object_or_404(AboutFeature, id=feature_id)
        feature.delete()
        messages.success(request, 'Đã xóa tiêu chí nổi bật!')
    return redirect('custom_manager')
# ==========================================================

# 9. Tra cứu lịch hẹn
def tra_cuu(request):
    if request.user.is_authenticated:
        bookings = Booking.objects.filter(user=request.user).order_by('-id')
        return render(request, 'tra_cuu.html', {'bookings': bookings, 'is_auto_history': True})
    else:
        phone_query = request.GET.get('phone')
        bookings = None
        if phone_query:
            bookings = Booking.objects.filter(phone=phone_query).order_by('-id')
        return render(request, 'tra_cuu.html', {'bookings': bookings, 'phone_query': phone_query})


# 1. HÀM XUẤT EXCEL
def export_products_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sach vay cuoi"

    # Tạo dòng tiêu đề cột
    columns = ['ID', 'Tên Váy', 'Giá thuê', 'Số lượng', 'Mô tả ngắn', 'Mô tả chi tiết', 'Chi nhánh', 'Trạng thái']
    ws.append(columns)

    products = Product.objects.all()
    for p in products:
        ws.append([
            p.id,
            p.name,
            p.price,
            p.quantity, 
            p.short_description,
            p.description,
            p.store.name if p.store else "",
            p.get_status_display()
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Danh_sach_vay_cuoi.xlsx"'
    wb.save(response)
    return response

# 2. HÀM NHẬP EXCEL
def import_products_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        
        # Bắt lỗi nếu file không phải Excel
        if not excel_file.name.endswith('.xlsx'):
            messages.error(request, "❌ Lỗi: Vui lòng tải lên file Excel (.xlsx)")
            return redirect('custom_manager')

        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            count = 0
            # Duyệt từ dòng 2 (bỏ qua dòng 1 là tiêu đề)
            for row in ws.iter_rows(min_row=2, values_only=True):
                # Ánh xạ đúng với số cột của file xuất ra ở trên:
                # row[0]: ID (bỏ qua)
                # row[1]: Tên Váy
                # row[2]: Giá thuê
                # row[3]: Số lượng
                # row[4]: Mô tả ngắn
                # row[5]: Mô tả chi tiết
                # row[6]: Chi nhánh
                
                name = row[1]
                if not name: 
                    continue # Nếu tên trống thì bỏ qua dòng đó luôn
                    
                price = row[2] if row[2] else 0
                quantity = row[3] if row[3] else 1
                short_desc = row[4] if row[4] else ""
                full_desc = row[5] if row[5] else ""
                store_name = row[6]
                
                store = None
                if store_name:
                    # Dùng icontains để tìm tên chi nhánh cho linh hoạt (gõ hoa thường đều được)
                    store = Store.objects.filter(name__icontains=store_name).first()

                Product.objects.create(
                    name=name,
                    price=price,
                    quantity=quantity,
                    short_description=short_desc,
                    description=full_desc,
                    store=store,
                    status='available' # Váy mới nhập kho mặc định là Sẵn sàng
                )
                count += 1
            
            messages.success(request, f"✅ Đã nhập thành công {count} váy cưới từ Excel!")
        except Exception as e:
            messages.error(request, f"❌ Lỗi khi đọc file: {str(e)}")
            
    return redirect('custom_manager')

# 12. Hàm thêm váy vào danh sách thử (Giỏ hàng)
def add_to_trial(request, product_id):
    trial_list = request.session.get('trial_list', [])
    if product_id not in trial_list:
        trial_list.append(product_id)
        request.session['trial_list'] = trial_list
        messages.success(request, "🛒 Đã thêm váy vào danh sách chờ thử thành công!")
    else:
        messages.info(request, "Váy này đã có trong danh sách của bạn rồi nhé.")
    return redirect('product_detail', product_id=product_id)

# 13. Hàm xử lý giỏ hàng
def trial_cart(request):
    trial_list = request.session.get('trial_list', [])
    products = Product.objects.filter(id__in=trial_list)

    if request.method == 'POST':
        if not request.user.is_authenticated:
            messages.warning(request, "Vui lòng đăng nhập để đặt lịch!")
            return redirect('login')

        name = request.POST.get('customer_name')
        phone_input = request.POST.get('phone_number') 
        date = request.POST.get('booking_date')
        note = request.POST.get('note')
        is_shipping = request.POST.get('is_shipping') 
        address = request.POST.get('address') if is_shipping else ''
        shipping_fee = request.POST.get('shipping_fee') if is_shipping else 0

        if name and phone_input and date and products.exists():
            new_booking = Booking.objects.create(
                user=request.user,
                customer_name=name, 
                phone=phone_input,  
                booking_date=date, 
                note=note,
                address=address,             
                shipping_fee=shipping_fee
            )

            for p in products:
                new_booking.products.add(p)

            list_vay = ", ".join([p.name for p in products]) 
            tieu_de_mail = f"[ĐƠN MỚI] {name} vừa đặt lịch thử {products.count()} váy"
            noi_dung_mail = (
                f"Chào Admin,\n\n"
                f"Có một đơn đặt lịch thử váy mới từ hệ thống:\n"
                f"- Khách hàng: {name}\n"
                f"- SĐT: {phone_input}\n"
                f"- Danh sách váy: {list_vay}\n"
                f"- Hình thức: {'Giao tận nhà' if is_shipping else 'Thử tại tiệm'}\n"
                f"- Ngày hẹn: {date}\n"
                f"- Ghi chú: {note}\n\n"
                f"Vui lòng vào trang quản trị để xử lý đơn hàng!"
            )

            try:
                send_mail(
                    subject=tieu_de_mail,
                    message=noi_dung_mail,
                    from_email='system@bridalluxury.com',
                    recipient_list=['admin_test@gmail.com'], 
                    fail_silently=True,
                )
            except Exception as e:
                print(f"Lỗi gửi mail: {e}")

            request.session['trial_list'] = []
            messages.success(request, f"✨ Đã chốt lịch thử {products.count()} váy thành công!")
            return redirect('tra_cuu') 
            
        else:
            messages.error(request, "Vui lòng điền đủ thông tin hoặc giỏ hàng đang trống.")

    return render(request, 'trial_cart.html', {'products': products})

# 14. Hàm xoá váy khỏi giỏ
def remove_from_trial(request, product_id):
    trial_list = request.session.get('trial_list', [])
    if product_id in trial_list:
        trial_list.remove(product_id)
        request.session['trial_list'] = trial_list
        messages.success(request, "🗑️ Đã bỏ váy khỏi danh sách thử.")
    return redirect('trial_cart') 

# 15. KÍCH HOẠT TÀI KHOẢN QUA EMAIL
def activate(request, uidb64, token):
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except(TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None

    if user is not None and default_token_generator.check_token(user, token):
        user.is_active = True 
        user.save()
        messages.success(request, '🎉 Chúc mừng! Tài khoản của bạn đã được kích hoạt thành công. Đăng nhập ngay!')
        return redirect('login')
    else:
        messages.error(request, '❌ Link kích hoạt không hợp lệ hoặc đã hết hạn!')
        return redirect('home')
    
# 16. Trang quản lý người dùng (Admin)
# Hàm Khóa / Mở khóa tài khoản khách hàng
def toggle_user_status(request, user_id):
    if request.user.is_staff:
        user = get_object_or_404(User, pk=user_id)
        # Chống admin tự khóa chính mình
        if user.is_superuser:
            messages.error(request, "Không thể khóa tài khoản Admin tối cao!")
        else:
            user.is_active = not user.is_active # Đảo ngược trạng thái
            user.save()
            trang_thai = "Mở khóa" if user.is_active else "Khóa"
            messages.success(request, f"Đã {trang_thai} tài khoản '{user.username}'!")
    return redirect('custom_manager')

# Hàm Xóa tài khoản
def delete_user(request, user_id):
    if request.user.is_staff:
        user = get_object_or_404(User, pk=user_id)
        if user.is_superuser:
            messages.error(request, "Không thể xóa tài khoản Admin tối cao!")
        else:
            user.delete()
            messages.success(request, f"Đã xóa vĩnh viễn tài khoản '{user.username}'!")
    return redirect('custom_manager')